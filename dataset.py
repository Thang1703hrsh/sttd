import pandas as pd
import numpy as np 

import streamlit as st 
from streamlit_option_menu import option_menu 

import os 
import streamlit_authenticator as stauth
import datetime 
import database as db 
import openpyxl 

# file_path = r'Line Balancing.xlsm'
# input_data = pd.read_excel(file_path, sheet_name='cap_ultra_test',skiprows=3,usecols='B:H')

path = 'Line Balancing_SE.xlsm'
from openpyxl import load_workbook
wb = load_workbook(filename = path)
g_sheet=wb.sheetnames
print(g_sheet)
# for i in g_sheet:
#     print(i)

# file modification timestamp of a file
m_time = os.path.getmtime(path)
# convert timestamp into DateTime object
dt_m = datetime.datetime.fromtimestamp(m_time)
print('Ngày tạo file :', dt_m)
list_dt_m = [dt_m]*len(g_sheet)
# file creation timestamp in float
c_time = os.path.getctime(path)
# convert creation timestamp into DateTime object
dt_c = datetime.datetime.fromtimestamp(c_time)
list_dt_c = [dt_c]*len(g_sheet)
print('Ngày chỉnh sửa file :', dt_c)
name_hat = ['Ultra Adventure hat']*len(g_sheet)
cat_hat = ['SA']*len(g_sheet)
data = {'Tên kiểu nón': name_hat ,'Tên loại hàng': cat_hat ,'Tên file': g_sheet , 'Ngày tạo file': list_dt_m , 'Ngày chỉnh sửa file': list_dt_c}

df = pd.DataFrame (data)

print(g_sheet)


def add_stream_url(track_ids):
	return [f'https://open.spotify.com/track/{t}' for t in track_ids]

def make_clickable(url, text):
    return f'<a target="_blank" href="{url}">{text}</a>'

# show data
df['Thông tin dữ liệu'] = add_stream_url(df['Tên file'])
df['Thông tin dữ liệu'] = df['Thông tin dữ liệu'].apply(make_clickable, args = (f'Infor' , ))
st.write(df.to_html(escape = False), unsafe_allow_html = True)

